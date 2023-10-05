import pyodbc
from openpyxl import load_workbook
import shutil
import os

# Define your MS SQL Server connection details
server = '192.168.100.90'
database = 'DataCollection'
username = 'sa'
password = 'Cannon45!'

# Establish the connection
conn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+password)

# Define your SQL query
sql_query = """
select serial_no, data_matrix, label_id from laser_result_entry
where is_deleted = 0
"""

# Execute the SQL query and fetch data
cursor = conn.cursor()
cursor.execute(sql_query)
query_result = cursor.fetchall()

# Load the Excel template
template_path = r'C:\excel_import\laser_result_template.xlsx'
if not os.path.exists(template_path):
    print(f"Error: Excel template '{template_path}' not found.")
else:
    # Create a new copy of the template
    new_excel_path =  r'C:\excel_import\new_file.xlsx'
    shutil.copyfile(template_path, new_excel_path)

    # Load the new copy of the Excel workbook
    workbook = load_workbook(new_excel_path)

    # Select the worksheet in the workbook (you may need to specify the sheet name)
    worksheet = workbook.active

    # Start populating data in row 2 (A2, B2, C2, etc.) for each row of query result
    row_number = 2  # Start from row 2
    for row_data in query_result:
        # Column indices (0-based) for each column in the row
        column1_index = 0  # A2
        column2_index = 1  # B2
        column3_index = 2  # C2

        # Insert data into specific cell coordinates for the current row
        worksheet.cell(row=row_number, column=column1_index + 1, value=row_data[0])  # A2
        worksheet.cell(row=row_number, column=column2_index + 1, value=row_data[1])  # B2
        worksheet.cell(row=row_number, column=column3_index + 1, value=row_data[2])  # C2

        # Move to the next row
        row_number += 1

    # Save the modified workbook with data to a new file
    workbook.save(new_excel_path)

    print(f"Data has been populated in '{new_excel_path}'")
