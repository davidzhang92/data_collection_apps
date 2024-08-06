from flask import Flask, request, jsonify, Response
import pyodbc
import datetime
from io import BytesIO
from openpyxl import load_workbook  # Import openpyxl
from openpyxl.styles import Alignment
from datetime import date
from secret_key import SECRET_KEY
import jwt
from functools import wraps

app = Flask(__name__)

# # Define your MS SQL Server connection details (Windows)
# server = '192.168.100.121'
# database = 'DataCollection'
# username = 'sa'
# password = 'Cannon45!'

# # Establish the connection
# conn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)


# def fetch_data(query):
#     cursor = conn.cursor()
#     cursor.execute(query)
#     rows = cursor.fetchall()
#     return rows


# Define your MS SQL Server connection details (Linux)
# Use the DSN you've defined in your odbc.ini file

dsn = 'DataCollection'

# Establish the connection
conn = pyodbc.connect('DSN=DataCollection;UID=sa;PWD=Cannon45!')

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        access_token = None

        
        if 'Authorization' in request.headers:
            access_token = request.headers['Authorization']


        if 'x-access-token' in request.headers:
            access_token = request.headers['x-access-token']

        if not access_token:
            return jsonify({'message': 'Session timed out, please login again.'}), 401

        try:
            data = jwt.decode(access_token, SECRET_KEY, algorithms=['HS256'])
            access_level = data.get('access_level')

            # if access_level not in ['read-only', 'operator', 'admin']:
            if access_level not in ['read-only', 'operator', 'admin']:
                return jsonify({'message': 'Error: You don\'t have sufficient privilege to perform this action.'}), 403
            
        except Exception as e:
            print(e)
            return jsonify({'message': 'Session timed out, please login again.', 'error': str(e)}), 401

        return f(*args, **kwargs)

    return decorated

# --------------------
# API endpoint get all programming result entry
# --------------------
@app.route('/api/get_programming_result_entry_view', methods=['GET'])
@token_required
def get_programming_result_entry_view():
    try:
        # Extract the page query parameter from the request
        page = int(request.args.get('page', 1))  # Default value is 1

        # Calculate the offset based on the page number
        offset = (page - 1) * 10

        cursor = conn.cursor()

        # Construct the SQL query to select data from the part_master table with OFFSET
        query = """SELECT 
                        b.part_no,
                        b.part_description,
                        serial_no, 
                        result, 
                        a.fail_current, 
                        a.fail_hr, 
                        a.fail_pairing, 
                        a.fail_bluetooth, 
                        a.fail_sleep_mode, 
                        a.fail_other, 
						d.defect_no,
						d.defect_description,
                        c.username,
                        a.created_date 
                    FROM programming_result_entry a 
                    INNER JOIN part_master b ON a.part_id = b.id 
                    LEFT JOIN user_master c ON a.created_by = c.id
					LEFT JOIN defect_master d on a.defect_id = d.id
					order by a.created_date desc
                    OFFSET ? ROWS FETCH NEXT 10 ROWS ONLY"""

        cursor.execute(query, (offset,))
        rows = cursor.fetchall()

        # Convert the result into a list of dictionaries for JSON serialization
        results = []
        columns = [column[0] for column in cursor.description]

        for row in rows:
            results.append(dict(zip(columns, row)))

        return jsonify(results)

    except Exception as e:
        return jsonify({'message': 'Error occurred while fetching data.', 'error': str(e)}), 500
    
# --------------------
# API endpoint filtering after clicking Search button
# --------------------

@app.route('/api/get_filter_search_programming_result_entry_view', methods=['GET'])
@token_required
def get_filter_search_programming_result_entry_view():
        
    selected_part_no = request.args.get('search_part_no')  # Get the selected value from the query parameters
    selected_date_from = request.args.get('search_date_from')  # Get the selected value from the query parameters
    selected_date_to = request.args.get('search_date_to')  # Get the selected value from the query parameters

    # if selected_part_no is None:
    #     return jsonify({'error': 'Search parameters not provided'})
    

    cursor = conn.cursor()
    
    # Construct the SQL query to select all data from the programming result entry table
    query = """SELECT 
                    b.part_no,
                    b.part_description,
                    serial_no, 
                    result, 
                    a.fail_current, 
                    a.fail_hr, 
                    a.fail_pairing, 
                    a.fail_bluetooth, 
                    a.fail_sleep_mode, 
                    a.fail_other, 
                    d.defect_no,
                    d.defect_description,
                    c.username,
                    a.created_date 
                FROM programming_result_entry a 
                INNER JOIN part_master b ON a.part_id = b.id 
                LEFT JOIN user_master c ON a.created_by = c.id
                LEFT JOIN defect_master d on a.defect_id = d.id
                WHERE 1=1"""

    parameters = []

    if selected_part_no:
        query += " AND b.part_no LIKE ?"
        parameters.append(f"%{selected_part_no}%")

    if selected_date_from:
        # Append the time '00:00:00' to the date
        date_from_with_time = f"{selected_date_from} 00:00:00"
        query += " AND a.created_date >= ?"
        parameters.append(datetime.datetime.strptime(date_from_with_time, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S'))


    if selected_date_to:
        # Append the time '23:59:59' to the date
        date_to_with_time = f"{selected_date_to} 23:59:59"
        query += " AND a.created_date <= ?"
        parameters.append(datetime.datetime.strptime(date_to_with_time, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S'))


    query += "  AND a.is_deleted = '0' ORDER BY a.created_date DESC;"

    # Construct the parameter values with wildcards
    cursor.execute(query, parameters)
    rows = cursor.fetchall()
    print(selected_date_from)
    print(selected_date_to)
    # Convert the result into a list of dictionaries for JSON serialization
    results = []
    columns = [column[0] for column in cursor.description]
    
    for row in rows:
        results.append(dict(zip(columns, row)))
    
    return jsonify(results)


# --------------------
# API endpoint download report
# --------------------
@app.route('/api/get_programming_result_report', methods=['POST'])
@token_required
def get_programming_result_report():
    # Define the path to the Excel template (modify this path accordingly)
    template_path = r'/data-storage/sfdc_apps/excel_import/programming_report_template.xlsx'

    try:
        # Try to parse JSON data from the request body
        request_data = request.get_json()

        # Check if request_data is None or not a dictionary
        if request_data is None or not isinstance(request_data, dict):
            return jsonify({'error': 'Invalid JSON data in the request body.'}), 400

        # Get the search parameters from the JSON data
        selected_part_no = request_data.get('part_no')
        selected_date_from = request_data.get('date_from')
        selected_date_to = request_data.get('date_to')
        user_generator = request_data.get('user_id')

        # Check if required parameters are provided
        # if not selected_part_id and not selected_date_from and not selected_date_to:
        #     return jsonify({'error': 'Missing required parameters. Please provide at least one search parameter.'}), 400

        cursor = conn.cursor()



        # Construct the SQL query to select data based on the provided parameters
        # Construct for Header report data
        query_header_data = """SELECT DISTINCT 
                                b.part_no,
                                b.part_description
                            FROM programming_result_entry a 
                            INNER JOIN part_master b ON a.part_id = b.id """
        parameters_header_data = []
        

        if selected_part_no:
            query_header_data += "WHERE b.part_no LIKE ?"
            parameters_header_data.append('%' + selected_part_no + '%')
        else: 

            if selected_date_from:
                query_header_data += "WHERE a.created_date >= ?"
                parameters_header_data.append(selected_date_from)
            elif selected_date_to:
                query_header_data += "WHERE a.created_date <= ?"
                parameters_header_data.append(selected_date_to)
            elif selected_date_from and selected_date_to:
                query_header_data += "WHERE a.created_date >= ? AND a.created_date <= ?"
                parameters_header_data.append(selected_date_from)
                parameters_header_data.append(selected_date_to)

        query_header_data += " AND a.is_deleted = '0' "

        # Execute the SQL query and fetch data
        cursor.execute(query_header_data, parameters_header_data)
        query_header_result = cursor.fetchall()

        # query who is the generator
        query_generator = "select username from user_master where id = ?"

        # Construct SQL for report data
        query_data = """SELECT 
                        b.part_no,
                        b.part_description,
                        serial_no, 
                        result, 
                        d.defect_no,
                        CONCAT(
                            d.defect_description,
                            CASE WHEN a.fail_current = 1 THEN ' current' ELSE '' END,
                            CASE WHEN a.fail_hr = 1 THEN ' hr' ELSE '' END,
                            CASE WHEN a.fail_pairing = 1 THEN 'pairing' ELSE '' END,
                            CASE WHEN a.fail_bluetooth = 1 THEN ' bluetooth' ELSE '' END,
                            CASE WHEN a.fail_sleep_mode = 1 THEN ' sleep mode' ELSE '' END,
                            CASE WHEN a.fail_other = 1 THEN ' other' ELSE '' END
                        ) AS defect_description,
                        c.username,
                        a.created_date 
                    FROM programming_result_entry a 
                    INNER JOIN part_master b ON a.part_id = b.id 
                    LEFT JOIN user_master c ON a.created_by = c.id
                    LEFT JOIN defect_master d on d.id = a.defect_id
                    """
        parameters_data = []

        conditions = []
        if selected_part_no:
            conditions.append("b.part_no LIKE ?")
            parameters_data.append('%' + selected_part_no + '%')

        if selected_date_from:
            date_from_with_time = f"{selected_date_from} 00:00:00"
            conditions.append("a.created_date >= ?")
            parameters_data.append(datetime.datetime.strptime(date_from_with_time, '%Y-%m-%d %H:%M:%S'))

        if selected_date_to:
            date_to_with_time = f"{selected_date_to} 23:59:59"
            conditions.append("a.created_date <= ?")
            parameters_data.append(datetime.datetime.strptime(date_to_with_time, '%Y-%m-%d %H:%M:%S'))

        conditions.append("a.is_deleted = '0'")

        if conditions:
            query_data += " WHERE " + " AND ".join(conditions)

        query_data += " ORDER BY a.created_date DESC;"

        # Execute the SQL query and fetch data
        cursor.execute(query_data, parameters_data)
        query_result = cursor.fetchall()

        # Execute the SQL query and fetch the generator
        generated_by = None  # Define generated_by before the if statement
        cursor.execute(query_generator, user_generator)
        generator_result = cursor.fetchone()
        if generator_result is not None:
            generated_by, = generator_result

        # Load the Excel template
        workbook = load_workbook(template_path)
        worksheet = workbook.active

        ########for header data##########
        # Data population
        part_no_data = [data[0] for data in query_header_result]
        part_description_data = [data[1] for data in query_header_result]

        # Join the data with ;
        part_no_joined = ';'.join(part_no_data)
        part_description_joined = ';'.join(part_description_data)

        # Unmerge cells
        # worksheet.unmerge_cells('C5')
        # worksheet.unmerge_cells('C6')

        # Write the data into the cells
        worksheet['C5'] = part_no_joined
        worksheet['C6'] = part_description_joined
        worksheet['C3'] = selected_date_from
        worksheet['C4'] = selected_date_to
        worksheet['I6'] = generated_by
        worksheet['I5'] = date.today().strftime('%Y-%m-%d')



        # Merge cells again and adjust the allignment
        worksheet.merge_cells('C5:F5')
        worksheet.merge_cells('C6:F6')
        worksheet['C5'].alignment = Alignment(horizontal='left')
        worksheet['C6'].alignment = Alignment(horizontal='left')

        ###########for data report#########
        row_number = 9  # Start from row 9

        for row_data in query_result:
           
            column2_index = 2  # Column B
            column3_index = 3  # Column C
            column4_index = 4  # Column D
            column5_index = 5 
            column6_index = 6 
            column7_index = 7 
            column8_index = 8 
            column9_index = 9 



            worksheet.cell(row=row_number, column=column2_index, value=row_data[0]) 
            worksheet.cell(row=row_number, column=column3_index, value=row_data[1]) 
            worksheet.cell(row=row_number, column=column4_index, value=row_data[2]) 
            worksheet.cell(row=row_number, column=column5_index, value=row_data[3]) 
            worksheet.cell(row=row_number, column=column6_index, value=row_data[4]) 
            worksheet.cell(row=row_number, column=column7_index, value=row_data[5]) 
            worksheet.cell(row=row_number, column=column8_index, value=row_data[6]) 
            worksheet.cell(row=row_number, column=column9_index, value=row_data[7]) 
 


            row_number += 1

        # Save the modified workbook in memory
        excel_output = BytesIO()
        workbook.save(excel_output)
        excel_output.seek(0)

        # Send the modified Excel file as a binary attachment
        return Response(
            excel_output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=programming_report.xlsx'
            }
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run()
    