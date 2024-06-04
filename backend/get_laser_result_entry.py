from flask import Flask, request, jsonify, Response
import pyodbc
import datetime
from io import BytesIO
from openpyxl import load_workbook  # Import openpyxl
from secret_key import SECRET_KEY
import jwt
from functools import wraps

app = Flask(__name__)

# # Define your MS SQL Server connection details (Windows)
# server = '192.168.100.121'
# database = 'DataCollection'
# username = 'sa'
# password = 'Cannon45!'

# # Establish the connection
# conn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)


# def fetch_data(query):
#     cursor = conn.cursor()
#     cursor.execute(query)
#     rows = cursor.fetchall()
#     return rows


# Define your MS SQL Server connection details (Linux)
# Use the DSN you've defined in your odbc.ini file

dsn = 'DataCollection'

# Establish the connection
conn = pyodbc.connect('DSN=DataCollection;UID=sa;PWD=Cannon45!')

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        access_token = None

        
        if 'Authorization' in request.headers:
            access_token = request.headers['Authorization']


        if 'x-access-token' in request.headers:
            access_token = request.headers['x-access-token']

        if not access_token:
            return jsonify({'message': 'Session timed out, please login again.'}), 401

        try:
            data = jwt.decode(access_token, SECRET_KEY, algorithms=['HS256'])
            access_level = data.get('access_level')

            # if access_level not in ['read-only', 'operator', 'admin']:
            if access_level not in ['read-only', 'operator', 'admin']:
                return jsonify({'message': 'Error: You don\'t have sufficient privilege to perform this action.'}), 403
            
        except Exception as e:
            print(e)
            return jsonify({'message': 'Session timed out, please login again.', 'error': str(e)}), 401

        return f(*args, **kwargs)

    return decorated
# --------------------
# API endpoint get all leaktest result entry
# --------------------
@app.route('/api/get_laser_result_entry', methods=['POST'])
@token_required
def get_laser_result_entry():
    # Define the path to the Excel template (modify this path accordingly)
    template_path = r'/data-storage/sfdc_apps/excel_import/laser_result_template.xlsx'

    try:
        # Try to parse JSON data from the request body
        request_data = request.get_json()

        # Check if request_data is None or not a dictionary
        if request_data is None or not isinstance(request_data, dict):
            return jsonify({'error': 'Invalid JSON data in the request body.'}), 400

        # Get the search parameters from the JSON data
        selected_part_id = request_data.get('part_id')
        selected_date_from = request_data.get('date_from')
        selected_date_to = request_data.get('date_to')
        selected_wo_no = request_data.get('wo_no')

        # Check if required parameters are provided
        if not selected_part_id and not selected_date_from and not selected_date_to:
            return jsonify({'error': 'Missing required parameters. Please provide at least one search parameter.'}), 400

        cursor = conn.cursor()

        # Construct the SQL query to select data based on the provided parameters
        query = """SELECT 
                    a.serial_no, 
                    a.data_matrix, 
                    a.label_id,
					a.wo_no,
					c.username

                FROM 
                    laser_result_entry a 
					
                INNER JOIN part_master b ON a.part_id = b.id 
				INNER JOIN user_master c ON a.created_by = c.id 
                WHERE wo_no = ?"""
        parameters = [selected_wo_no]

        if selected_part_id:
            query += " AND a.part_id = ?"
            parameters.append(selected_part_id)

        if selected_date_from:
            date_from_with_time = f"{selected_date_from}"
            query += " AND a.created_date >= ?"
            parameters.append(datetime.datetime.strptime(date_from_with_time, '%Y/%m/%d %H:%M'))

        if selected_date_to:
            date_to_with_time = f"{selected_date_to}"
            query += " AND a.created_date <= ?"
            parameters.append(datetime.datetime.strptime(date_to_with_time, '%Y/%m/%d %H:%M'))

        query += "  AND a.is_deleted = '0' ORDER BY a.created_date DESC;"

        # Execute the SQL query and fetch data
        cursor.execute(query, parameters)
        query_result = cursor.fetchall()

        # if query_result is None:
        #      return jsonify({'error : no data found'}), 500

  # Load the Excel template
        workbook = load_workbook(template_path)
        worksheet = workbook.active


      
        row_number = 2  # Start from row 2

        for row_data in query_result:
           
            column1_index = 1  # Column A
            column2_index = 2  # Column B
            column3_index = 3 # Column C
            column4_index = 5 
            column5_index = 6   

            worksheet.cell(row=row_number, column=column1_index, value=row_data[0])  
            worksheet.cell(row=row_number, column=column2_index, value=row_data[1]) 
            worksheet.cell(row=row_number, column=column3_index, value=row_data[2])
            worksheet.cell(row=row_number, column=column4_index, value=row_data[3])
            worksheet.cell(row=row_number, column=column5_index, value=row_data[4]) 

            row_number += 1

        # Save the modified workbook in memory
        excel_output = BytesIO()
        workbook.save(excel_output)
        excel_output.seek(0)

        # Send the modified Excel file as a binary attachment
        return Response(
            excel_output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=laser_result.xlsx'
            }
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run()