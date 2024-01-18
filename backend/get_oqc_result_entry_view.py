from flask import Flask, request, jsonify, Response
import pyodbc
import datetime
from io import BytesIO
from openpyxl import load_workbook  # Import openpyxl
from openpyxl.styles import Alignment
from datetime import date

app = Flask(__name__)

# # Define your MS SQL Server connection details (Windows)
# server = '192.168.100.121'
# database = 'DataCollection'
# username = 'sa'
# password = 'Cannon45!'

# # Establish the connection
# conn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)


# def fetch_data(query):
#     cursor = conn.cursor()
#     cursor.execute(query)
#     rows = cursor.fetchall()
#     return rows


# Define your MS SQL Server connection details (Linux)
# Use the DSN you've defined in your odbc.ini file

dsn = 'DataCollection'

# Establish the connection
conn = pyodbc.connect('DSN=DataCollection;UID=sa;PWD=Cannon45!')


# --------------------
# API endpoint get all leaktest result entry
# --------------------
@app.route('/api/get_oqc_result_entry_view', methods=['GET'])
def get_oqc_result_entry_view():
    try:
        # Extract the page query parameter from the request
        page = int(request.args.get('page', 1))  # Default value is 1

        # Calculate the offset based on the page number
        offset = (page - 1) * 10

        cursor = conn.cursor()

        # Construct the SQL query to select data from the part_master table with OFFSET
        query = "select a.id as id, b.part_no, c.defect_description, a.result, a.serial_no, a.created_date from oqc_result_entry a inner join part_master b on a.part_id = b.id left join defect_master c on a.defect_id = c.id where a.is_deleted='0' order by created_date desc OFFSET ? ROWS FETCH NEXT 10 ROWS ONLY"

        cursor.execute(query, (offset,))
        rows = cursor.fetchall()

        # Convert the result into a list of dictionaries for JSON serialization
        results = []
        columns = [column[0] for column in cursor.description]

        for row in rows:
            results.append(dict(zip(columns, row)))

        return jsonify(results)

    except Exception as e:
        return jsonify({'message': 'Error occurred while fetching data.', 'error': str(e)}), 500
    
# --------------------
# API endpoint filtering after clicking Search button
# --------------------

@app.route('/api/get_filter_search_oqc_result_entry_view', methods=['GET'])
def get_filter_search_oqc_result_entry_view():
        
    selected_part_no = request.args.get('search_part_no')  # Get the selected value from the query parameters
    selected_date_from = request.args.get('search_date_from')  # Get the selected value from the query parameters
    selected_date_to = request.args.get('search_date_to')  # Get the selected value from the query parameters

    # if selected_part_no is None:
    #     return jsonify({'error': 'Search parameters not provided'})
    

    cursor = conn.cursor()
    
    # Construct the SQL query to select all data from the leaktest result entry table
    query = "select a.id as id, b.part_no, c.defect_description, a.result, a.serial_no, a.created_date from oqc_result_entry a inner join part_master b on a.part_id = b.id left join defect_master c on a.defect_id = c.id WHERE 1=1"

    parameters = []

    if selected_part_no:
        query += " AND b.part_no like ?"
        parameters.append(f"%{selected_part_no}%")

    if selected_date_from:
        # Append the time '00:00:00' to the date
        date_from_with_time = f"{selected_date_from} 00:00:00"
        query += " AND a.created_date >= ?"
        parameters.append(datetime.datetime.strptime(date_from_with_time, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S'))


    if selected_date_to:
        # Append the time '23:59:59' to the date
        date_to_with_time = f"{selected_date_to} 23:59:59"
        query += " AND a.created_date <= ?"
        parameters.append(datetime.datetime.strptime(date_to_with_time, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S'))


    query += "  AND a.is_deleted = '0' ORDER BY a.created_date DESC;"

    # Construct the parameter values with wildcards
    cursor.execute(query, parameters)
    rows = cursor.fetchall()
    print(selected_date_from)
    print(selected_date_to)
    # Convert the result into a list of dictionaries for JSON serialization
    results = []
    columns = [column[0] for column in cursor.description]
    
    for row in rows:
        results.append(dict(zip(columns, row)))
    
    return jsonify(results)

    # --------------------
# API endpoint download report
# --------------------
@app.route('/api/get_oqc_result_report', methods=['POST'])
def get_oqc_result_report():
    # Define the path to the Excel template (modify this path accordingly)
    template_path = r'/data-storage/sfdc_apps/excel_import/oqc_report_template.xlsx'

    try:
        # Try to parse JSON data from the request body
        request_data = request.get_json()

        # Check if request_data is None or not a dictionary
        if request_data is None or not isinstance(request_data, dict):
            return jsonify({'error': 'Invalid JSON data in the request body.'}), 400

        # Get the search parameters from the JSON data
        selected_part_no = request_data.get('part_no')
        selected_date_from = request_data.get('date_from')
        selected_date_to = request_data.get('date_to')

        # Check if required parameters are provided
        # if not selected_part_id and not selected_date_from and not selected_date_to:
        #     return jsonify({'error': 'Missing required parameters. Please provide at least one search parameter.'}), 400

        cursor = conn.cursor()



        # Construct the SQL query to select data based on the provided parameters
        # Construct for Header report data
        query_header_data = """SELECT DISTINCT 
                                b.part_no,
                                b.part_description
                            FROM oqc_result_entry a 
                            INNER JOIN part_master b ON a.part_id = b.id """
        parameters_header_data = []
        

        if selected_part_no:
            query_header_data += "WHERE b.part_no LIKE ?"
            parameters_header_data.append('%' + selected_part_no + '%')

        query_header_data += " AND a.is_deleted = '0' "

        # Execute the SQL query and fetch data
        cursor.execute(query_header_data, parameters_header_data)
        query_header_result = cursor.fetchall()


        # Construct SQL for report data
        query_data = """SELECT 
                        b.part_no,
                        b.part_description,
                        serial_no, 
                        result, 
                        c.defect_description,
                        a.created_date 
                    FROM oqc_result_entry a 
                    INNER JOIN part_master b ON a.part_id = b.id
                    LEFT JOIN defect_master c ON a.defect_id = c.id"""
        parameters_data = []

        conditions = []
        if selected_part_no:
            conditions.append("b.part_no LIKE ?")
            parameters_data.append('%' + selected_part_no + '%')

        if selected_date_from:
            date_from_with_time = f"{selected_date_from} 00:00:00"
            conditions.append("a.created_date >= ?")
            parameters_data.append(datetime.datetime.strptime(date_from_with_time, '%Y-%m-%d %H:%M:%S'))

        if selected_date_to:
            date_to_with_time = f"{selected_date_to} 23:59:59"
            conditions.append("a.created_date <= ?")
            parameters_data.append(datetime.datetime.strptime(date_to_with_time, '%Y-%m-%d %H:%M:%S'))

        conditions.append("a.is_deleted = '0'")

        if conditions:
            query_data += " WHERE " + " AND ".join(conditions)

        query_data += " ORDER BY a.created_date DESC;"

        # Execute the SQL query and fetch data
        cursor.execute(query_data, parameters_data)
        query_result = cursor.fetchall()

        # Load the Excel template
        workbook = load_workbook(template_path)
        worksheet = workbook.active

        ########for header data##########
        # Data population
        part_no_data = [data[0] for data in query_header_result]
        part_description_data = [data[1] for data in query_header_result]

        # Join the data with ;
        part_no_joined = ';'.join(part_no_data)
        part_description_joined = ';'.join(part_description_data)

        # Unmerge cells
        # worksheet.unmerge_cells('C5')
        # worksheet.unmerge_cells('C6')

        # Write the data into the cells
        worksheet['C5'] = part_no_joined
        worksheet['C6'] = part_description_joined
        worksheet['G5'] = date.today().strftime('%Y-%m-%d')
        worksheet['C3'] = selected_date_from
        worksheet['C4'] = selected_date_to

        # Merge cells again and adjust the allignment
        worksheet.merge_cells('C5:E5')
        worksheet.merge_cells('C6:E6')
        worksheet['C5'].alignment = Alignment(horizontal='left')
        worksheet['C6'].alignment = Alignment(horizontal='left')

        ###########for data report#########
        row_number = 9  # Start from row 9

        for row_data in query_result:
           
            column2_index = 2  # Column B
            column3_index = 3  # Column C
            column4_index = 4  # Column D
            column5_index = 5 
            column6_index = 6 
            column7_index = 7 
 


            worksheet.cell(row=row_number, column=column2_index, value=row_data[0]) 
            worksheet.cell(row=row_number, column=column3_index, value=row_data[1]) 
            worksheet.cell(row=row_number, column=column4_index, value=row_data[2]) 
            worksheet.cell(row=row_number, column=column5_index, value=row_data[3]) 
            worksheet.cell(row=row_number, column=column6_index, value=row_data[4]) 
            worksheet.cell(row=row_number, column=column7_index, value=row_data[5]) 


            row_number += 1

        # Save the modified workbook in memory
        excel_output = BytesIO()
        workbook.save(excel_output)
        excel_output.seek(0)

        # Send the modified Excel file as a binary attachment
        return Response(
            excel_output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=oqc_report.xlsx'
            }
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


    
if __name__ == '__main__':
    app.run()
    